import feedparser
import pandas as pd
from datetime import datetime, timedelta
import re
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Date filter ──────────────────────────────────────────────────────────────
CUTOFF = datetime.now() - timedelta(days=365)

# ── Company list ─────────────────────────────────────────────────────────────
COMPANIES = [
    # Issuers – Major Banks
    {"name": "Chase",            "segment": "Issuers – Major Banks",         "hq": "New York, US",      "founded": 1799, "website": "chase.com"},
    {"name": "Bank of America",  "segment": "Issuers – Major Banks",         "hq": "Charlotte, US",     "founded": 1998, "website": "bankofamerica.com"},
    {"name": "Citi",             "segment": "Issuers – Major Banks",         "hq": "New York, US",      "founded": 1812, "website": "citi.com"},
    {"name": "Wells Fargo",      "segment": "Issuers – Major Banks",         "hq": "San Francisco, US", "founded": 1852, "website": "wellsfargo.com"},
    {"name": "Capital One",      "segment": "Issuers – Major Banks",         "hq": "McLean, US",        "founded": 1994, "website": "capitalone.com"},
    {"name": "Discover",         "segment": "Issuers – Major Banks",         "hq": "Riverwoods, US",    "founded": 1985, "website": "discover.com"},
    {"name": "American Express", "segment": "Issuers – Major Banks",         "hq": "New York, US",      "founded": 1850, "website": "americanexpress.com"},
    {"name": "Synchrony",        "segment": "Issuers – Major Banks",         "hq": "Stamford, US",      "founded": 2003, "website": "synchrony.com"},
    {"name": "Barclays US",      "segment": "Issuers – Major Banks",         "hq": "Wilmington, US",    "founded": 1736, "website": "barclaysus.com"},
    # Issuers – Fintech / Neobank Cards
    {"name": "Apple Card",       "segment": "Issuers – Fintech/Neobank",     "hq": "Cupertino, US",     "founded": 2019, "website": "apple.com/apple-card"},
    {"name": "Chime",            "segment": "Issuers – Fintech/Neobank",     "hq": "San Francisco, US", "founded": 2013, "website": "chime.com"},
    {"name": "Current",          "segment": "Issuers – Fintech/Neobank",     "hq": "New York, US",      "founded": 2015, "website": "current.com"},
    {"name": "Varo",             "segment": "Issuers – Fintech/Neobank",     "hq": "San Francisco, US", "founded": 2015, "website": "varomoney.com"},
    {"name": "Dave",             "segment": "Issuers – Fintech/Neobank",     "hq": "Los Angeles, US",   "founded": 2016, "website": "dave.com"},
    {"name": "Petal",            "segment": "Issuers – Fintech/Neobank",     "hq": "New York, US",      "founded": 2016, "website": "petalcard.com"},
    {"name": "Deserve",          "segment": "Issuers – Fintech/Neobank",     "hq": "Menlo Park, US",    "founded": 2013, "website": "deserve.com"},
    {"name": "Step",             "segment": "Issuers – Fintech/Neobank",     "hq": "San Francisco, US", "founded": 2018, "website": "step.com"},
    {"name": "Brex",             "segment": "Issuers – Fintech/Neobank",     "hq": "San Francisco, US", "founded": 2017, "website": "brex.com"},
    # Credit Card Networks
    {"name": "Visa",             "segment": "Credit Card Networks",          "hq": "San Francisco, US", "founded": 1958, "website": "visa.com"},
    {"name": "Mastercard",       "segment": "Credit Card Networks",          "hq": "Purchase, US",      "founded": 1966, "website": "mastercard.com"},
    # BNPL & Installment Credit
    {"name": "Affirm",           "segment": "BNPL & Installment Credit",     "hq": "San Francisco, US", "founded": 2012, "website": "affirm.com"},
    {"name": "Klarna",           "segment": "BNPL & Installment Credit",     "hq": "Stockholm, Sweden", "founded": 2005, "website": "klarna.com"},
    {"name": "Afterpay",         "segment": "BNPL & Installment Credit",     "hq": "Melbourne, AU",     "founded": 2014, "website": "afterpay.com"},
    {"name": "Sezzle",           "segment": "BNPL & Installment Credit",     "hq": "Minneapolis, US",   "founded": 2016, "website": "sezzle.com"},
    {"name": "Splitit",          "segment": "BNPL & Installment Credit",     "hq": "New York, US",      "founded": 2012, "website": "splitit.com"},
    {"name": "Zip",              "segment": "BNPL & Installment Credit",     "hq": "Sydney, AU",        "founded": 2013, "website": "zip.co"},
    {"name": "PayPal Pay Later", "segment": "BNPL & Installment Credit",     "hq": "San Jose, US",      "founded": 1998, "website": "paypal.com"},
    # Credit Building & Underwriting
    {"name": "Experian Boost",   "segment": "Credit Building & Underwriting","hq": "Dublin, Ireland",   "founded": 1996, "website": "experian.com"},
    {"name": "Credit Karma",     "segment": "Credit Building & Underwriting","hq": "San Francisco, US", "founded": 2007, "website": "creditkarma.com"},
    {"name": "Self Financial",   "segment": "Credit Building & Underwriting","hq": "Austin, US",        "founded": 2015, "website": "self.inc"},
    {"name": "Kikoff",           "segment": "Credit Building & Underwriting","hq": "San Francisco, US", "founded": 2019, "website": "kikoff.com"},
    {"name": "Grow Credit",      "segment": "Credit Building & Underwriting","hq": "Santa Monica, US",  "founded": 2018, "website": "growcredit.com"},
    {"name": "Perch",            "segment": "Credit Building & Underwriting","hq": "New York, US",      "founded": 2020, "website": "perchcredit.com"},
    # Mortgage Originators – Traditional
    {"name": "Rocket Mortgage",  "segment": "Mortgage Originators – Traditional","hq": "Detroit, US",   "founded": 1985, "website": "rocketmortgage.com"},
    {"name": "UWM",              "segment": "Mortgage Originators – Traditional","hq": "Pontiac, US",    "founded": 1986, "website": "uwm.com"},
    {"name": "loanDepot",        "segment": "Mortgage Originators – Traditional","hq": "Foothill Ranch, US","founded": 2010, "website": "loandepot.com"},
    {"name": "Fairway Mortgage", "segment": "Mortgage Originators – Traditional","hq": "Madison, US",   "founded": 1996, "website": "fairwayindependentmortgage.com"},
    {"name": "Guild Mortgage",   "segment": "Mortgage Originators – Traditional","hq": "San Diego, US", "founded": 1960, "website": "guildmortgage.com"},
    {"name": "PennyMac",         "segment": "Mortgage Originators – Traditional","hq": "Westlake Village, US","founded": 2008,"website": "pennymac.com"},
    # Mortgage Originators – Fintech
    {"name": "Better.com",       "segment": "Mortgage Originators – Fintech","hq": "New York, US",      "founded": 2014, "website": "better.com"},
    {"name": "Morty",            "segment": "Mortgage Originators – Fintech","hq": "New York, US",      "founded": 2016, "website": "morty.com"},
    {"name": "Tomo",             "segment": "Mortgage Originators – Fintech","hq": "Stamford, US",      "founded": 2020, "website": "tomo.com"},
    {"name": "Flyhomes",         "segment": "Mortgage Originators – Fintech","hq": "Seattle, US",       "founded": 2015, "website": "flyhomes.com"},
    {"name": "Orchard",          "segment": "Mortgage Originators – Fintech","hq": "New York, US",      "founded": 2017, "website": "orchard.com"},
    # Mortgage Infrastructure & Tech
    {"name": "Blend",            "segment": "Mortgage Infrastructure & Tech","hq": "San Francisco, US", "founded": 2012, "website": "blend.com"},
    {"name": "ICE Mortgage",     "segment": "Mortgage Infrastructure & Tech","hq": "Atlanta, US",       "founded": 1997, "website": "icemortgage.com"},
    {"name": "Maxwell",          "segment": "Mortgage Infrastructure & Tech","hq": "Denver, US",        "founded": 2015, "website": "himaxwell.com"},
    {"name": "nCino",            "segment": "Mortgage Infrastructure & Tech","hq": "Wilmington, US",    "founded": 2012, "website": "ncino.com"},
    {"name": "Polly",            "segment": "Mortgage Infrastructure & Tech","hq": "San Francisco, US", "founded": 2019, "website": "polly.io"},
    {"name": "Optimal Blue",     "segment": "Mortgage Infrastructure & Tech","hq": "Plano, US",         "founded": 2002, "website": "optimalblue.com"},
    # Mortgage Servicers & Secondary Market
    {"name": "Mr. Cooper",       "segment": "Mortgage Servicers & Secondary","hq": "Coppell, US",       "founded": 1994, "website": "mrcooper.com"},
    {"name": "Pennymac",         "segment": "Mortgage Servicers & Secondary","hq": "Westlake Village, US","founded": 2008,"website": "pennymac.com"},
    {"name": "Fannie Mae",       "segment": "Mortgage Servicers & Secondary","hq": "Washington, US",    "founded": 1938, "website": "fanniemae.com"},
    {"name": "Freddie Mac",      "segment": "Mortgage Servicers & Secondary","hq": "McLean, US",        "founded": 1970, "website": "freddiemac.com"},
    {"name": "Ginnie Mae",       "segment": "Mortgage Servicers & Secondary","hq": "Washington, US",    "founded": 1968, "website": "ginniemae.gov"},
    # Regulatory Bodies
    {"name": "CFPB",             "segment": "Regulatory Bodies",             "hq": "Washington, US",    "founded": 2011, "website": "consumerfinance.gov"},
    {"name": "OCC",              "segment": "Regulatory Bodies",             "hq": "Washington, US",    "founded": 1863, "website": "occ.gov"},
    {"name": "FHFA",             "segment": "Regulatory Bodies",             "hq": "Washington, US",    "founded": 2008, "website": "fhfa.gov"},
    {"name": "HUD",              "segment": "Regulatory Bodies",             "hq": "Washington, US",    "founded": 1965, "website": "hud.gov"},
    {"name": "FTC",              "segment": "Regulatory Bodies",             "hq": "Washington, US",    "founded": 1914, "website": "ftc.gov"},
    {"name": "Federal Reserve",  "segment": "Regulatory Bodies",             "hq": "Washington, US",    "founded": 1913, "website": "federalreserve.gov"},
]

# ── Topic classification keywords ─────────────────────────────────────────────
TOPIC_RULES = {
    "Funding":          r"fund|raise|series [a-e]|ipo|valuation|invest|capital raise|round",
    "M&A":              r"\bacquir|\bmerger\b|\bdeal\b|\bbuy\b|takeover|acquisition",
    "Partnerships":     r"partner|collaborat|integrat|joint venture|team(s)? up|agreement",
    "Regulatory":       r"regulat|compliance|enforcement|fine|penalty|lawsuit|cfpb|occ|fhfa|hud|ftc|federal reserve|settlement|consent order",
    "Product":          r"launch|new product|feature|app|update|introduc|announc",
    "Rates":            r"\bapr\b|\brate\b|basis point|interest rate|prime rate|mortgage rate|yield",
    "Rewards":          r"reward|cashback|cash back|points|miles|bonus|sign.?up offer|loyalty|perks",
    "Disruptions":      r"lawsuit|breach|hack|outage|layoff|job cut|downgrad|loss|scandal|exec.*depart|ceo.*resign|fraud|default|delinquency|chargeback|fine|penalt|negative|downtur|fail|bankrupt|recall|complaint",
    "Global Expansion": r"expand|international|global|launch.*country|launch.*market|europe|asia|latin america|africa|canada|uk|australia|india|brazil|mexico|singapore|japan",
    "Competition":      r"compet|rival|market share|versus|vs\.|head.to.head|challeng|disrupt|underprice|outpace|beat|overtake|new entrant",
}

REGIONS = {
    "Europe":        r"europe|uk|germany|france|spain|italy|netherlands|sweden|denmark|ireland|poland",
    "Asia-Pacific":  r"asia|australia|japan|singapore|india|china|korea|hong kong|new zealand",
    "Latin America": r"latin america|brazil|mexico|colombia|argentina|chile|peru",
    "Africa":        r"africa|nigeria|kenya|south africa|ghana|egypt",
    "Canada":        r"canada|canadian",
    "Middle East":   r"middle east|uae|saudi|dubai|israel|qatar",
    "North America": r"united states|us market|american market",
}

DISRUPTION_SEVERITY = {
    "High":   r"bankrupt|fraud|major breach|class.action|billion.*fine|mass layoff|federal.*charges|criminal",
    "Medium": r"lawsuit|breach|hack|layoff|job cut|ceo.*resign|exec.*depart|regulat.*action|consent order|settlement",
    "Low":    r"complaint|downgrad|minor outage|negative|loss|chargeback|delinquency",
}

# ── Regex extractors ──────────────────────────────────────────────────────────
def extract_amount(text):
    m = re.search(r'\$[\d,.]+\s*(million|billion|M|B)\b', text, re.I)
    return m.group(0) if m else ""

def extract_round(text):
    m = re.search(r'Series\s+[A-F]|Seed|Pre.?Seed|IPO|SPAC|debt round|growth round', text, re.I)
    return m.group(0) if m else ""

def extract_valuation(text):
    m = re.search(r'valued?\s+at\s+\$[\d,.]+\s*(billion|million)', text, re.I)
    return m.group(0) if m else ""

def extract_investors(text):
    m = re.search(r'led by ([A-Z][^\.,]+)', text)
    return m.group(1) if m else ""

def extract_partner(text):
    m = re.search(r'partner(?:s|ed|ing)? with ([A-Z][^\.,]+)', text, re.I)
    return m.group(1) if m else ""

def extract_rate(text):
    m = re.search(r'(\d+\.?\d*)\s*%\s*(APR|interest|mortgage rate|prime)?', text, re.I)
    return m.group(0) if m else ""

def extract_reward(text):
    patterns = [
        r'\d+\.?\d*x?\s*(?:cash\s*back|points|miles)',
        r'\$[\d,]+\s+(?:bonus|sign.?up)',
        r'\d+%\s+(?:cash\s*back|rewards)',
    ]
    for p in patterns:
        m = re.search(p, text, re.I)
        if m:
            return m.group(0)
    return ""

def classify_topic(text):
    text_l = text.lower()
    for topic, pattern in TOPIC_RULES.items():
        if re.search(pattern, text_l):
            return topic
    return "General"

def classify_region(text):
    text_l = text.lower()
    for region, pattern in REGIONS.items():
        if re.search(pattern, text_l):
            return region
    return "Global"

def classify_severity(text):
    text_l = text.lower()
    for severity, pattern in DISRUPTION_SEVERITY.items():
        if re.search(pattern, text_l):
            return severity
    return "Low"

def extract_competitor(text):
    m = re.search(r'(?:vs\.?|versus|rival|compete[sd]? with|challenge[sd]?)\s+([A-Z][a-zA-Z\s]+?)(?:,|\.|\s+in\s|\s+on\s)', text)
    return m.group(1).strip() if m else ""

# ── RSS fetch ─────────────────────────────────────────────────────────────────
def fetch_news(company_name):
    query = company_name.replace(" ", "+") + "+credit+card+OR+mortgage+OR+loan+OR+finance"
    url = f"https://news.google.com/rss/search?q={query}&hl=en-US&gl=US&ceid=US:en"
    articles = []
    try:
        feed = feedparser.parse(url)
        for entry in feed.entries:
            pub = entry.get("published_parsed")
            if not pub:
                continue
            pub_dt = datetime(*pub[:6])
            if pub_dt < CUTOFF:
                continue
            articles.append({
                "title":    entry.get("title", ""),
                "url":      entry.get("link", ""),
                "date":     pub_dt.strftime("%Y-%m-%d"),
                "source":   entry.get("source", {}).get("title", ""),
            })
        time.sleep(0.3)
    except Exception as e:
        print(f"  ⚠ Error fetching {company_name}: {e}")
    return articles

# ── Main scrape ───────────────────────────────────────────────────────────────
def scrape_all():
    rows_news, rows_financials, rows_partnerships = [], [], []
    rows_rates, rows_rewards, rows_disruptions = [], [], []
    rows_expansion, rows_competition = [], []

    total = len(COMPANIES)
    for i, co in enumerate(COMPANIES, 1):
        name = co["name"]
        segment = co["segment"]
        print(f"[{i}/{total}] {name}")
        articles = fetch_news(name)

        for a in articles:
            title = a["title"]
            url   = a["url"]
            date  = a["date"]
            topic = classify_topic(title)

            rows_news.append({
                "Company": name, "Segment": segment, "Date": date,
                "Topic": topic, "Headline": title, "Source": a["source"], "URL": url,
            })

            if topic == "Funding":
                rows_financials.append({
                    "Company": name, "Segment": segment, "Date": date,
                    "Round": extract_round(title), "Amount": extract_amount(title),
                    "Valuation": extract_valuation(title), "Investors": extract_investors(title),
                    "Headline": title, "URL": url,
                })

            if topic == "Partnerships":
                rows_partnerships.append({
                    "Company": name, "Segment": segment, "Date": date,
                    "Partner": extract_partner(title),
                    "Headline": title, "URL": url,
                })

            if topic == "Rates":
                rows_rates.append({
                    "Company": name, "Segment": segment, "Date": date,
                    "Product Type": "Mortgage" if re.search(r'mortgage|home loan|refi', title, re.I) else "Credit Card",
                    "Rate Mentioned": extract_rate(title),
                    "Headline": title, "URL": url,
                })

            if topic == "Rewards":
                rows_rewards.append({
                    "Company": name, "Segment": segment, "Date": date,
                    "Reward Type": (
                        "Miles" if re.search(r'miles', title, re.I) else
                        "Points" if re.search(r'points', title, re.I) else
                        "Cashback" if re.search(r'cash.?back', title, re.I) else
                        "Sign-up Bonus" if re.search(r'sign.?up|bonus', title, re.I) else "Other"
                    ),
                    "Reward Mentioned": extract_reward(title),
                    "Headline": title, "URL": url,
                })

            if topic == "Disruptions":
                rows_disruptions.append({
                    "Company": name, "Segment": segment, "Date": date,
                    "Disruption Type": (
                        "Data Breach/Hack" if re.search(r'breach|hack', title, re.I) else
                        "Lawsuit/Legal" if re.search(r'lawsuit|suit|charges|litigation', title, re.I) else
                        "Layoffs" if re.search(r'layoff|job cut', title, re.I) else
                        "Leadership" if re.search(r'ceo|exec.*depart|resign', title, re.I) else
                        "Financial Loss" if re.search(r'loss|default|delinquency|bankrupt', title, re.I) else
                        "Regulatory Action" if re.search(r'fine|penalt|enforcement|consent order', title, re.I) else
                        "Outage/Service" if re.search(r'outage|down|fail', title, re.I) else "Other"
                    ),
                    "Severity": classify_severity(title),
                    "Headline": title, "URL": url,
                })

            if topic == "Global Expansion":
                rows_expansion.append({
                    "Company": name, "Segment": segment, "Date": date,
                    "Region": classify_region(title),
                    "Entry Type": (
                        "Acquisition" if re.search(r'acquir|merger', title, re.I) else
                        "Partnership" if re.search(r'partner', title, re.I) else
                        "Regulatory Approval" if re.search(r'licens|approv|regulat', title, re.I) else
                        "Product Launch" if re.search(r'launch|introduc', title, re.I) else "Expansion"
                    ),
                    "Headline": title, "URL": url,
                })

            if topic == "Competition":
                rows_competition.append({
                    "Company": name, "Segment": segment, "Date": date,
                    "Competitor": extract_competitor(title),
                    "Event Type": (
                        "New Entrant" if re.search(r'new entrant|enter.*market', title, re.I) else
                        "Pricing" if re.search(r'underpric|lower rate|fee cut', title, re.I) else
                        "Feature Parity" if re.search(r'feature|product.*match|copy', title, re.I) else
                        "Market Share" if re.search(r'market share|overtake|beat', title, re.I) else "Competitive Move"
                    ),
                    "Headline": title, "URL": url,
                })

    return {
        "news": rows_news, "financials": rows_financials, "partnerships": rows_partnerships,
        "rates": rows_rates, "rewards": rows_rewards, "disruptions": rows_disruptions,
        "expansion": rows_expansion, "competition": rows_competition,
    }

# ── Excel styling helpers ─────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", start_color="1F3864")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
ALT_FILL      = PatternFill("solid", start_color="EAF0FB")
BORDER_SIDE   = Side(style="thin", color="CCCCCC")
CELL_BORDER   = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)

SEGMENT_COLORS = {
    "Issuers – Major Banks":              "D6E4F7",
    "Issuers – Fintech/Neobank":          "D6F7E4",
    "Credit Card Networks":               "F7F0D6",
    "BNPL & Installment Credit":          "F7D6D6",
    "Credit Building & Underwriting":     "EBD6F7",
    "Mortgage Originators – Traditional": "D6F7F7",
    "Mortgage Originators – Fintech":     "F7E8D6",
    "Mortgage Infrastructure & Tech":     "F7D6F0",
    "Mortgage Servicers & Secondary":     "E8F7D6",
    "Regulatory Bodies":                  "F7F7D6",
}

SEVERITY_COLORS = {"High": "FF4444", "Medium": "FFA500", "Low": "FFD700"}

def style_sheet(ws, headers, col_widths, rows, color_col=None, color_map=None):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(c)].width = col_widths[c - 1]
    ws.row_dimensions[1].height = 28

    for r, row in enumerate(rows, 2):
        fill = ALT_FILL if r % 2 == 0 else PatternFill()
        for c, key in enumerate(headers, 1):
            val = row.get(key, "")
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = CELL_BORDER
            if color_col and color_map and key == color_col:
                hex_c = color_map.get(val, "")
                if hex_c:
                    cell.fill = PatternFill("solid", start_color=hex_c)
                else:
                    cell.fill = fill
            else:
                cell.fill = fill

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

# ── Summary sheet ─────────────────────────────────────────────────────────────
def write_summary(wb, data):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False

    title_font = Font(bold=True, size=14, color="1F3864", name="Arial")
    sub_font   = Font(bold=True, size=10, name="Arial")
    val_font   = Font(size=10, name="Arial")

    ws["B2"] = "Consumer Finance Market Intelligence Tracker"
    ws["B2"].font = title_font
    ws["B3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["B3"].font = Font(italic=True, size=9, name="Arial", color="666666")
    ws["B4"] = f"Coverage: Last 12 months  |  Companies tracked: {len(COMPANIES)}"
    ws["B4"].font = Font(size=9, name="Arial", color="666666")

    stats = [
        ("Total Articles",       len(data["news"])),
        ("Funding Events",       len(data["financials"])),
        ("Partnerships",         len(data["partnerships"])),
        ("Rate Events",          len(data["rates"])),
        ("Rewards Events",       len(data["rewards"])),
        ("Disruptions",          len(data["disruptions"])),
        ("Global Expansion",     len(data["expansion"])),
        ("Competition Events",   len(data["competition"])),
    ]

    ws["B6"] = "Metric"
    ws["C6"] = "Count"
    ws["B6"].font = sub_font
    ws["C6"].font = sub_font

    for i, (label, count) in enumerate(stats, 7):
        ws.cell(row=i, column=2, value=label).font = val_font
        ws.cell(row=i, column=3, value=count).font = Font(bold=True, size=10, name="Arial")

    ws["B16"] = "Articles by Segment"
    ws["B16"].font = sub_font
    from collections import Counter
    seg_counts = Counter(a["Segment"] for a in data["news"])
    for i, (seg, cnt) in enumerate(sorted(seg_counts.items(), key=lambda x: -x[1]), 17):
        ws.cell(row=i, column=2, value=seg).font = val_font
        ws.cell(row=i, column=3, value=cnt).font = val_font

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 14

# ── Build workbook ────────────────────────────────────────────────────────────
def build_excel(data, path):
    wb = Workbook()
    wb.remove(wb.active)

    write_summary(wb, data)

    # Companies
    ws_co = wb.create_sheet("Companies")
    co_headers = ["Name", "Segment", "HQ", "Founded", "Website"]
    co_widths   = [22, 34, 22, 10, 30]
    co_rows = [{
        "Name": c["name"], "Segment": c["segment"], "HQ": c["hq"],
        "Founded": c["founded"], "Website": c["website"]
    } for c in COMPANIES]
    style_sheet(ws_co, co_headers, co_widths, co_rows, color_col="Segment", color_map=SEGMENT_COLORS)

    # News
    ws_news = wb.create_sheet("News")
    news_headers = ["Date", "Company", "Segment", "Topic", "Headline", "Source", "URL"]
    news_widths   = [12, 20, 30, 16, 60, 20, 50]
    style_sheet(ws_news, news_headers, news_widths,
                sorted(data["news"], key=lambda x: x["Date"], reverse=True))

    # Financials
    ws_fin = wb.create_sheet("Financials")
    fin_headers = ["Date", "Company", "Segment", "Round", "Amount", "Valuation", "Investors", "Headline", "URL"]
    fin_widths   = [12, 20, 30, 14, 16, 22, 30, 55, 45]
    style_sheet(ws_fin, fin_headers, fin_widths,
                sorted(data["financials"], key=lambda x: x["Date"], reverse=True))

    # Partnerships
    ws_part = wb.create_sheet("Partnerships")
    part_headers = ["Date", "Company", "Segment", "Partner", "Headline", "URL"]
    part_widths   = [12, 20, 30, 28, 60, 45]
    style_sheet(ws_part, part_headers, part_widths,
                sorted(data["partnerships"], key=lambda x: x["Date"], reverse=True))

    # Rates
    ws_rates = wb.create_sheet("Rates")
    rate_headers = ["Date", "Company", "Segment", "Product Type", "Rate Mentioned", "Headline", "URL"]
    rate_widths   = [12, 20, 30, 16, 18, 60, 45]
    style_sheet(ws_rates, rate_headers, rate_widths,
                sorted(data["rates"], key=lambda x: x["Date"], reverse=True))

    # Rewards
    ws_rew = wb.create_sheet("Rewards")
    rew_headers = ["Date", "Company", "Segment", "Reward Type", "Reward Mentioned", "Headline", "URL"]
    rew_widths   = [12, 20, 30, 16, 22, 60, 45]
    style_sheet(ws_rew, rew_headers, rew_widths,
                sorted(data["rewards"], key=lambda x: x["Date"], reverse=True))

    # Disruptions
    ws_dis = wb.create_sheet("Disruptions")
    dis_headers = ["Date", "Company", "Segment", "Disruption Type", "Severity", "Headline", "URL"]
    dis_widths   = [12, 20, 30, 22, 10, 60, 45]
    style_sheet(ws_dis, dis_headers, dis_widths,
                sorted(data["disruptions"], key=lambda x: (x["Severity"], x["Date"]), reverse=True),
                color_col="Severity", color_map=SEVERITY_COLORS)

    # Global Expansion
    ws_exp = wb.create_sheet("Global Expansion")
    exp_headers = ["Date", "Company", "Segment", "Region", "Entry Type", "Headline", "URL"]
    exp_widths   = [12, 20, 30, 18, 20, 60, 45]
    style_sheet(ws_exp, exp_headers, exp_widths,
                sorted(data["expansion"], key=lambda x: x["Date"], reverse=True))

    # Competition
    ws_comp = wb.create_sheet("Competition")
    comp_headers = ["Date", "Company", "Segment", "Competitor", "Event Type", "Headline", "URL"]
    comp_widths   = [12, 20, 30, 24, 20, 60, 45]
    style_sheet(ws_comp, comp_headers, comp_widths,
                sorted(data["competition"], key=lambda x: x["Date"], reverse=True))

    wb.save(path)
    print(f"\n✅ Saved: {path}")

# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("🔍 Scraping Google News RSS...")
    data = scrape_all()
    print("\n📊 Building Excel workbook...")
    build_excel(data, "consumer_finance_intel.xlsx")

    print("\n── Summary ──────────────────────────────")
    print(f"  Total articles:     {len(data['news'])}")
    print(f"  Funding events:     {len(data['financials'])}")
    print(f"  Partnerships:       {len(data['partnerships'])}")
    print(f"  Rate events:        {len(data['rates'])}")
    print(f"  Rewards events:     {len(data['rewards'])}")
    print(f"  Disruptions:        {len(data['disruptions'])}")
    print(f"  Global expansion:   {len(data['expansion'])}")
    print(f"  Competition events: {len(data['competition'])}")
